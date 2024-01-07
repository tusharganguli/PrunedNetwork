#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Jul  1 12:01:19 2022

@author: tushar
"""

import pandas as pd
import os
import numpy as np
from scipy.linalg import svd
from tensorflow import keras

# for writing to excel files
import openpyxl 
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import tempfile

class LogHandler:
    
    def __init__(self, log_dir, tensorboard_dir, prune_dir, 
                 model_dir, plot_dir): 
        logdir_name = "LogDir"
        log_dir = "../" + logdir_name + "/" + log_dir + "/" 
        log_dir = os.path.join(os.curdir,log_dir)
        if not os.path.exists(log_dir):
            os.makedirs(log_dir)
        
        
        self.prune_dir = log_dir + prune_dir
        if not os.path.exists(self.prune_dir):
            os.makedirs(self.prune_dir)
        
        self.tensorboard_dir = log_dir + tensorboard_dir
        if not os.path.exists(self.tensorboard_dir):
            os.makedirs(self.tensorboard_dir)
        
        self.model_dir = log_dir + model_dir
        if not os.path.exists(self.model_dir):
            os.makedirs(self.model_dir)
        
        self.plot_dir = log_dir + plot_dir
        if not os.path.exists(self.plot_dir):
            os.makedirs(self.plot_dir)
            
        self.timestamp = get_timestamp()
        
        self.log_filename = tempfile.TemporaryFile()
        
        # variables for svd
        self.sig_df = pd.DataFrame() #self.__create_svd_df()
        self.svd_plot_info = pd.DataFrame(columns=['cacc','tpp'])
        #self.svd_plots = gp.SVDPlots()
        
        # variables for logging in model_run
        self.df = self.__create_data_frame()
        
        
    def __del__(self):
        del self.sig_df
        del self.svd_plot_info
        #del self.svd_plots
        
        # variables for logging in model_run
        del self.df
        del self.tensorboard_dir
        del self.prune_dir
        del self.model_dir
        del self.plot_dir

    def __create_data_frame(self):
        
        df = pd.DataFrame(columns = ['Pruning Type',
                                     'No Of Epochs', 
                                     'Pct Pruning', 
                                     'Prune At Accuracy',
                                     'Training Accuracy',
                                     'Validation Accuracy',
                                     'Test Accuracy',
                                     'Training Loss',
                                     'Validation Loss',
                                     'Test Loss',
                                     'Top-1 Training Accuracy',
                                     'Top-1 Validation Accuracy',
                                     'Top-1 Test Accuracy',
                                     'Top-5 Training Accuracy',
                                     'Top-5 Validation Accuracy',
                                     'Top-5 Test Accuracy'
                                     ])

          
        return df
    
    def set_log_filename(self, log_filename):
        self.log_filename = log_filename
        
    def get_tensorboard_dir(self):
        tb_dir = self.tensorboard_dir + "/" + self.log_filename 
        if not os.path.exists(tb_dir):
            os.makedirs(tb_dir)
        return tb_dir
    
    def get_modelname(self):
        model_dir = self.model_dir + "/" + self.log_filename
        if not os.path.exists(model_dir):
            os.makedirs(model_dir)
        return model_dir
    
    def get_plot_dir(self):
        plot_dir = self.plot_dir + "/" + self.log_filename
        if not os.path.exists(plot_dir):
            os.makedirs(plot_dir)
        return plot_dir
    
    def log_data(self, pruning_type, history, eval_result, 
                   num_epochs, prune_pct=0, prune_at_acc=0):
        
        train_loss = history.history["loss"][1]
        #train_loss = sum(loss_lst)/len(loss_lst)
        train_acc = history.history["accuracy"][-1]
        #train_accuracy = sum(t_acc_lst)/len(t_acc_lst)
        val_loss = history.history["val_loss"][-1]
        #val_loss = sum(val_loss)/len(val_loss)
        val_acc = history.history["val_accuracy"][-1]
        #val_accuracy = sum(val_acc)/len(val_acc)
        test_loss = eval_result[0]
        test_acc = eval_result[1]
        
        training_top1 = history.history["top1"][-1]
        #training_top1 = sum(training_top1_lst)/len(training_top1_lst)
        training_top5 = history.history["top5"][-1]
        #training_top5 = sum(training_top5_lst)/len(training_top5_lst)
        val_top1 = history.history["val_top1"][-1]
        #val_top1 = sum(val_top1_lst)/len(val_top1_lst)
        val_top5 = history.history["val_top5"][-1]
        #val_top5 = sum(val_top5_lst)/len(val_top5_lst)
        test_top1 = eval_result[2]
        #test_top1 = sum(test_top1_lst)/len(test_top1_lst)
        test_top5 = eval_result[3]
        #test_top5 = sum(test_top5_lst)/len(test_top5_lst)
        
        log_data = [pruning_type, num_epochs,
                    prune_pct,
                    prune_at_acc,
                    train_acc, val_acc,
                    test_acc, train_loss,
                    val_loss, test_loss,
                    training_top1, val_top1, test_top1,
                    training_top5, val_top5, test_top5
                    ]
        
        df2 = pd.DataFrame([log_data], columns=list(self.df))
        self.df = pd.concat([self.df,df2], ignore_index=True)
    
    def __generate_avg( self, history_list, evaluate_list):
        
        train_loss = train_accuracy = val_loss = 0
        val_accuracy = test_loss = test_accuracy = 0
        
        train_loss = history_list[-1].history["loss"][-1]
        train_accuracy = history_list[-1].history["accuracy"][-1]
        val_loss = history_list[-1].history["val_loss"][-1]
        val_accuracy = history_list[-1].history["val_accuracy"][-1]
        test_loss = evaluate_list[-1][0]
        test_accuracy = evaluate_list[-1][1]
        
        return (train_loss,train_accuracy, 
                val_loss, val_accuracy, 
                test_loss, test_accuracy )

    def write_to_file(self, prune_filename=""):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = dataframe_to_rows(self.df)
        
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)    
        if prune_filename == "":
            filename = self.prune_dir + "/" + self.log_filename + ".xls"
        else:
            filename = self.prune_dir + "/" + prune_filename + ".xls"
        wb.save(filename)
        self.df = self.df.iloc[0:0]
    
    def log_single_run(self, log_filename):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = dataframe_to_rows(self.df)
        
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)    
        log_filename = self.prune_dir + "/" + log_filename + ".xls"
        wb.save(log_filename)
        
    def get_sv(self, model):
        from scipy.linalg import svd
        
        sig_df = pd.DataFrame()
        
        for layer in model.layers:
            if not isinstance(layer,keras.layers.Dense) and \
                not isinstance(layer,keras.layers.Conv2D):
                continue
            weights = layer.get_weights()
            """
            if "conv" in layer.name:
                dim = weights[0].shape
                filters = dim[-1]
                for idx in range(filters):
                    u,s,vt = svd(np.matrix(weights[0][:,:,:,idx]))
                    df_s = pd.DataFrame(s)
                    sig_df = pd.concat([sig_df,df_s], ignore_index=True, axis=1)
            """
            if "dense" in layer.name:
                u,s,vt = svd(weights[0])
                df_s = pd.DataFrame(s)
                sig_df = pd.concat([sig_df,df_s], ignore_index=True, axis=1)
            
        return sig_df

    def log_sv(self, model):
        sig_df = self.get_sv(model)
        self.sig_df = pd.concat([self.sig_df,sig_df], ignore_index=True, axis=1)
        
    def log_prune_info(self, total_prune_pct, curr_acc):
        
        info = [[curr_acc,total_prune_pct]]
        df_info = pd.DataFrame(info, columns=self.svd_plot_info.columns)
        #self.svd_plot_info = self.svd_plot_info.append(df_info)
        self.svd_plot_info = pd.concat([self.svd_plot_info,df_info])
        #self.svd_plots.PlotRatio(self.svd_df,layer_cnt, curr_acc, 
        #                         self.total_pruning_pct, prune_dir)

    def get_svd_details(self, model):
        layer_cnt = get_num_layers(model)
        return [self.sig_df,self.svd_plot_info, layer_cnt]
     
    def write_svd(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        rows = dataframe_to_rows(self.sig_df)
        
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                 ws.cell(row=r_idx, column=c_idx, value=value)    
        
        date = datetime.now().strftime("%Y%m%d-%H%M%S")
        filename = "./svd/" + date + ".xls"
        wb.save(filename)
        #self.sig_df.to_excel(filename)
        self.sig_df = self.sig_df.iloc[0:0]
        
    def reset_svd(self):
        del self.sig_df
        self.sig_df = pd.DataFrame()
        del self.svd_plot_info
        self.svd_plot_info = pd.DataFrame(columns=['cacc','tpp'])
        #del self.svd_plots
        #self.svd_plots = gp.SVDPlots()

def get_exp_decay_range(samples, a=1, b=0.2):
    # a, b: exponential decay parameter
    # N: number of samples 
    vals = a * (1-b) ** np.arange(samples)
    vals = (vals/np.sum(vals))*100
    return vals

def add_time_to_filename(filename=""):
    import time
    run_id = time.strftime("_%Y_%m_%d-%H_%M_%S")
    run_id = filename+run_id
    return run_id

def get_timestamp():
    import time
    return time.strftime("_%Y_%m_%d-%H_%M_%S")

def get_time():
    import time
    return time.time_ns()

def get_num_layers(model):
    layer_cnt = 0
    for layer in model.layers:
        if not isinstance(layer,keras.layers.Dense) or "dense" not in layer.name:
            continue
        layer_cnt += 1
    return layer_cnt

def write(df, filename):
    
    #if not os.path.exists(filename):
    #    os.makedirs(filename)
        
    wb = openpyxl.Workbook()
    ws = wb.active
    rows = dataframe_to_rows(df)
    
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
             ws.cell(row=r_idx, column=c_idx, value=value)    
    wb.save(filename) 
    

def get_svd(model):
    u,sv,vt = __compute_svd(model)
    return u,sv,vt

def __compute_svd(model):
    
    sv_df = pd.DataFrame()
    u_df = pd.DataFrame()
    vt_df = pd.DataFrame()
    
    for layer in model.layers:
        if not isinstance(layer,keras.layers.Dense):
            continue
        weights = layer.get_weights()
        
        u,s,vt = svd(weights[0], full_matrices=False)
        
        sv_df = pd.concat([sv_df, pd.DataFrame(s)], ignore_index=True, axis=1)
        u_df = pd.concat([u_df, pd.DataFrame(u)], ignore_index=True, axis=1)
        vt_df = pd.concat([vt_df, pd.DataFrame(vt)], ignore_index=True, axis=1)
    return u_df,sv_df,vt_df


def __compute_norm(M):
    """
    Computes Spectral,Frobenius and Nuclear norm

    Parameters
    ----------
    M : Matrix
        .

    Returns
    -------
    Spectral, Forbenious and Nuclear norm.

    """
    M_s = svd(M, full_matrices=False, compute_uv=False)
    M_spec_norm = M_s[0]
    M_frob_norm = np.sqrt(np.dot(M_s,M_s))
    M_nuc_norm = np.sum(M_s)
    
    return (M_spec_norm, M_frob_norm, M_nuc_norm)
    
def __compute_diff_norm(M1, M2):
    """
    Computes the difference of norms for both the matrix based on their
    singular values.

    Parameters
    ----------
    M1 : Matrix 
        First Matrix
    M2 : Matrix
        Second Matrix.

    Returns
    -------
    Difference of norm values.

    """
    M1_spec_norm, M1_frob_norm, M1_nuc_norm = __compute_norm(M1)
    M2_spec_norm, M2_frob_norm, M2_nuc_norm = __compute_norm(M2)
    
    diff_spec = M1_spec_norm-M2_spec_norm
    diff_frob = M1_frob_norm-M2_frob_norm
    diff_nuc = M1_nuc_norm-M2_nuc_norm
    
    return (diff_spec,diff_frob,diff_nuc)

def generate_low_rank_matrix(A, k):
    """

    Parameters
    ----------
    A : Matrix
        Full matrix of which we have to generate a low rank matrix.
        
    k : Int
        Dimension to which the matrix has to be reduced
    Returns
    -------
    Low rank matrix.

    """
    
    u,s,vt = svd(A, full_matrices=False)
    #u[:,k:] = 0
    #s[k:] = 0
    #S = np.diag(s)
    #vt[k:,:] = 0
    
    u_dim = np.shape(u)[0]
    vt_dim = np.shape(vt)[1]
    
    Ak = np.zeros((u_dim, vt_dim))
    
    for i in range(0,k):
        #uu = np.reshape(u[:,i],(u_dim,1))
        #vv = np.reshape(vt[i,:],(1,vt_dim))
        #Ak += s[i] * (uu*vv)
        Ak += s[i] * np.outer(u.T[i], vt[i])
    
    #tmp = np.dot(u, np.dot(S, vt))
    return Ak        

def generate_matrix_norms(std_model, pruned_model):
    num_layers = len(std_model.layers)
    df = pd.DataFrame(columns=["Spectral", "Frobenius","Nuclear"])
    
    for layer_id in range(0,num_layers):
        std_layer = std_model.layers[layer_id]
        pruned_layer = pruned_model.layers[layer_id]
        #tf_pruned_layer = tf_pruned_model.layers[layer_id]
        
        # checking only one of the models for dense layer is sufficient
        if not isinstance(std_layer,keras.layers.Dense):
            continue
        
        std_wts = std_layer.get_weights()[0]
        pruned_wts = pruned_layer.get_weights()[0]
        #tf_pruned_wts = tf_pruned_layer.get_weights()[0]
        
        #pruned_matrix_rank = np.linalg.matrix_rank(pruned_wts)    
        #Ak = generate_low_rank_matrix(std_wts,pruned_matrix_rank)
        #Tk = generate_low_rank_matrix(tf_pruned_wts,pruned_matrix_rank)
        
        data = pd.DataFrame([["" ,"" ,"" ]],columns=list(df), 
                            index=["Layer"+str(layer_id)])
        df = df.append(data)
        del data
        
        
        spec_A, frob_A, nuc_A = __compute_norm(std_wts)
        #spec_Ak, frob_Ak, nuc_Ak = __compute_norm(Ak)
        spec_B, frob_B, nuc_B = __compute_norm(pruned_wts)
        #spec_T, frob_T, nuc_T = __compute_norm(tf_pruned_wts)
        #spec_Tk, frob_Tk, nuc_Tk = __compute_norm(Tk)
        
        data = pd.DataFrame([[spec_A,frob_A,nuc_A], 
                             #[spec_Ak,frob_Ak,nuc_Ak],
                             [spec_B,frob_B,nuc_B],
                             #[spec_T, frob_T, nuc_T],
                             #[spec_Tk, frob_Tk, nuc_Tk]
                             ], 
                            columns=list(df), 
                            index=["||Std||", "||Pruned||"])
        
        df = df.append(data)
        del data
        """
        spec_A_Ak, frob_A_Ak, nuc_A_Ak = __compute_diff_norm(std_wts, Ak)
        spec_A_B, frob_A_B, nuc_A_B = __compute_diff_norm(std_wts, pruned_wts)
        #spec_Ak_B, frob_Ak_B, nuc_Ak_B = __compute_diff_norm(Ak, pruned_wts)
        
        #spec_A_T, frob_A_T, nuc_A_T = __compute_diff_norm(std_wts, tf_pruned_wts)
        #spec_Ak_Tf, frob_Ak_Tf, nuc_Ak_Tf = __compute_diff_norm(Ak, tf_pruned_wts)
        #spec_A_Tk, frob_A_Tk, nuc_A_Tk = __compute_diff_norm(std_wts, Tk)
        
        #data = pd.DataFrame([[spec_A_Ak,frob_A_Ak,nuc_A_Ak], 
        #                     [spec_A_B,frob_A_B,nuc_A_B],
        #                     #[spec_Ak_B,frob_Ak_B,nuc_Ak_B],
        #                     [spec_A_T, frob_A_T, nuc_A_T],
        #                     #[spec_Ak_Tf, frob_Ak_Tf, nuc_Ak_Tf],
        #                     [spec_A_Tk, frob_A_Tk, nuc_A_Tk],
        #                     ], 
        #                    columns=list(df), 
        #                    index=["||Std||-||Std_k||", "||Std||-||Freq||",
        #                           "||Std||-||Mag||","||Std||-||Mag_k||"])
        #df = df.append(data)
        del data
        
        #spec_A_Ak,frob_A_Ak,nuc_A_Ak = __compute_norm(std_wts-Ak)
        spec_A_B,frob_A_B,nuc_A_B = __compute_norm(std_wts-pruned_wts)
        #spec_Ak_B,frob_Ak_B,nuc_Ak_B = __compute_norm(Ak-pruned_wts)
        
        #spec_A_T,frob_A_T,nuc_A_T = __compute_norm(std_wts-tf_pruned_wts)
        #spec_Ak_Tf,frob_Ak_Tf,nuc_Ak_Tf = __compute_norm(Ak-tf_pruned_wts)
        #spec_A_Tk,frob_A_Tk,nuc_A_Tk = __compute_norm(std_wts-Tk)
        
        data = pd.DataFrame([#[spec_A_Ak,frob_A_Ak,nuc_A_Ak],
                             [spec_A_B,frob_A_B,nuc_A_B],
                             #[spec_Ak_B,frob_Ak_B,nuc_Ak_B],
                             [spec_A_T,frob_A_T,nuc_A_T],
                             #[spec_Ak_Tf,frob_Ak_Tf,nuc_Ak_Tf],
                             #[spec_A_Tk,frob_A_Tk,nuc_A_Tk],
                             ], 
                            columns=list(df), 
                            index=["||Std-Freq||", 
                                   "||Std-Mag||"])
        #df = df.append(data)
        del data
        """
    return df


def evaluate_low_rank_approx(self, model, pruned_model):
    """
    Generate low rank approximation for each layer of the model. 
    Set the layer weights accodingly and evaluate the model

    Parameters
    ----------
    model : Keras model obect
        Model for which we need to generate low rank approximation.

    Raises
    ------
    ValueError
        DESCRIPTION.

    Returns
    -------
    Int
        Test Accuracy.

    """
    num_layers = len(model.layers)
    
    for layer_id in range(0,num_layers):
        layer = model.layers[layer_id]
        pruned_layer = pruned_model.layers[layer_id]
        
        # checking only one of the models for dense layer is sufficient
        if not isinstance(layer,keras.layers.Dense):
            continue
        
        wts = layer.get_weights()
        pruned_wts = pruned_layer.get_weights()[0]
        
        pruned_matrix_rank = np.linalg.matrix_rank(pruned_wts)    
        Ak = generate_low_rank_matrix(wts[0],pruned_matrix_rank)
        wts[0] = Ak
        layer.set_weights(wts)
    
    eval_result = model.evaluate(self.test_img,self.test_labels)
    return eval_result

